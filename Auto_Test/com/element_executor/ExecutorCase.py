from openpyxl import workbook
from openpyxl import load_workbook
from Auto_Test.config.rw_config import ConfigParser
from Auto_Test.com.element_executor.ExecutorSingle import ExecutorSingleCase
from Auto_Test.com.element_executor.Parameter import Parameter
from Auto_Test.com.driver.driver import Driver
from Auto_Test.com.util.Log import Log
import time
import os
from os.path import *


class ExecutorCase:
    con = ConfigParser()
    single_case = ExecutorSingleCase()
    par_exe = Parameter()
    driver = Driver()
    path = 'test-output/'
    dir_time = time.strftime('%Y-%m-%d', time.localtime(time.time()))
    screen_path = path + 'img/' + dir_time + '/'
    log_path = path + 'logs/'
    if not exists(screen_path):
        os.makedirs(screen_path)
    if not exists(log_path):
        os.makedirs(log_path)
    log_info = Log(log_path + 'info.log', level='info')
    log_error = Log(log_path + 'error.log', level='error')

    def read_excel(self,config_path):
        case_url = self.con.get_config(config_path, 'Case', 'case_url')
        wb = load_workbook(case_url)
        # 循环遍历所有sheet
        sheets = wb.sheetnames
        for i in range(len(sheets)):
            sheet = wb[sheets[i]]
            print('第' + str(i + 1) + '个sheet: ' + sheet.title + '->>>')
            self.sheet_data(wb,sheet,config_path,case_url)
        wb.close()
        return True

    def sheet_data(self,wb,sheet,config_path,case_url):
        case_data = None
        for row in range(sheet.min_row + 1, sheet.max_row + 1):
            flag = True
            for col in range(1, sheet.max_column):
                case_num = sheet.cell(row, 1).value
                case_module = sheet.cell(row,2).value
                if case_num is not None:
                    case_data = sheet.cell(row, col).value
                    if sheet.cell(sheet.min_row, col).value == '测试步骤(element_name,type,loc_type,locator)':
                        #先进行登录操作
                        driver = self.driver.driver_initial(self.driver,config_path,self.log_info,self.log_error)
                        print('---准备执行用例：%s' % case_num +' 测试模块:%s'%case_module+'---'+' 对应元素：\n%s' % case_data)
                        list_element = case_data.strip('\n').split(';')
                        if self.single_case.element_executor(driver,case_module,list_element,config_path,self.screen_path,self.log_path,self.log_info,self.log_error):
                            if sheet.cell(row, col + 1).value == 'url':
                                url = self.par_exe.get_url(driver)
                                self.con.write_config(config_path, case_module, 'url', url)
                        else:
                            flag = False
                            self.log_error.logger.error('用例：'+case_num+'执行失败')
                        driver.quit()
                    if sheet.cell(sheet.min_row, col).value == '实际结果':
                        sheet.cell(row=row, column=col).value = str(flag)
                        wb.save(case_url)
                        print('保存结果成功~~~')
                else:
                    # 空的行不执行
                    break